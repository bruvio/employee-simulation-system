#!/Users/brunoviola/bruvio-tools/.venv/bin/python3

import argparse
from datetime import datetime
import json
from pathlib import Path

from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd

from logger import LOGGER
from app_paths import get_artifact_path, get_table_path


class DataExportSystem:
    """Comprehensive data export system for employee simulation results.

    Supports CSV, Excel, and JSON formats with enhanced formatting.
    """

    def __init__(self, output_dir=None):
        # Use centralized paths instead of hardcoded directory
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        LOGGER.info("Initialized data export system with centralized paths")

    def export_employee_population(self, population_data, format_types=None):
        """Export employee population data in multiple formats.

        Args:
            population_data: List of employee dictionaries or DataFrame
            format_types: List of formats ['csv', 'excel', 'json'] or None for all

        Returns:
            dict: Paths to exported files
        """
        if format_types is None:
            format_types = ["csv", "excel", "json"]

        # Convert to DataFrame if needed
        if isinstance(population_data, list):
            df = pd.DataFrame(population_data)
        else:
            df = population_data.copy()

        exported_files = {}
        base_filename = f"employee_population_{self.timestamp}"

        LOGGER.info(f"Exporting population data ({len(df)} employees) in formats: {format_types}")

        # CSV Export
        if "csv" in format_types:
            csv_path = get_table_path(f"{base_filename}.csv")
            df.to_csv(csv_path, index=False)
            exported_files["csv"] = csv_path
            LOGGER.info(f"CSV export completed: {csv_path}")

        # Excel Export with enhanced formatting
        if "excel" in format_types:
            excel_path = get_table_path(f"{base_filename}.xlsx")
            self._export_excel_population(df, excel_path)
            exported_files["excel"] = excel_path
            LOGGER.info(f"Excel export completed: {excel_path}")

        # JSON Export with metadata
        if "json" in format_types:
            json_path = get_artifact_path(f"{base_filename}.json")
            self._export_json_population(df, json_path)
            exported_files["json"] = json_path
            LOGGER.info(f"JSON export completed: {json_path}")

        return exported_files

    def export_simulation_results(self, simulation_data, format_types=None):
        """Export multi-cycle simulation results.

        Args:
            simulation_data: Dictionary with cycle results or DataFrame
            format_types: List of formats to export

        Returns:
            dict: Paths to exported files
        """
        if format_types is None:
            format_types = ["csv", "excel", "json"]

        exported_files = {}
        base_filename = f"simulation_results_{self.timestamp}"

        LOGGER.info(f"Exporting simulation results in formats: {format_types}")

        # Convert simulation data to DataFrame if needed
        if isinstance(simulation_data, dict):
            # Handle case where we have multiple DataFrames
            if "inequality_metrics" in simulation_data:
                df = simulation_data["inequality_metrics"]
            else:
                df = pd.DataFrame(simulation_data)
        else:
            df = simulation_data.copy()

        # CSV Export
        if "csv" in format_types:
            csv_path = get_table_path(f"{base_filename}_inequality.csv")
            df.to_csv(csv_path, index=False)
            exported_files["csv"] = csv_path
            LOGGER.info(f"Simulation CSV export completed: {csv_path}")

        # Excel Export with multiple sheets
        if "excel" in format_types:
            excel_path = get_table_path(f"{base_filename}.xlsx")
            self._export_excel_simulation(simulation_data, excel_path)
            exported_files["excel"] = excel_path
            LOGGER.info(f"Simulation Excel export completed: {excel_path}")

        # JSON Export
        if "json" in format_types:
            json_path = get_artifact_path(f"{base_filename}.json")
            self._export_json_simulation(simulation_data, json_path)
            exported_files["json"] = json_path
            LOGGER.info(f"Simulation JSON export completed: {json_path}")

        return exported_files

    def export_analysis_report(self, population_data, simulation_results, format_types=None):
        """Export comprehensive analysis report combining population and simulation data.

        Args:
            population_data: Employee population data
            simulation_results: Multi-cycle simulation results
            format_types: List of formats to export

        Returns:
            dict: Paths to exported files
        """
        if format_types is None:
            format_types = ["excel", "json"]

        exported_files = {}
        base_filename = f"comprehensive_analysis_{self.timestamp}"

        LOGGER.info("Generating comprehensive analysis report")

        # Generate summary statistics
        analysis_summary = self._generate_analysis_summary(population_data, simulation_results)

        # Excel Export with comprehensive dashboard
        if "excel" in format_types:
            excel_path = get_table_path(f"{base_filename}.xlsx")
            self._export_excel_analysis(population_data, simulation_results, analysis_summary, excel_path)
            exported_files["excel"] = excel_path
            LOGGER.info(f"Analysis Excel report completed: {excel_path}")

        # JSON Export with nested structure
        if "json" in format_types:
            json_path = get_artifact_path(f"{base_filename}.json")
            self._export_json_analysis(population_data, simulation_results, analysis_summary, json_path)
            exported_files["json"] = json_path
            LOGGER.info(f"Analysis JSON report completed: {json_path}")

        return exported_files

    def _export_excel_population(self, df, filepath):
        """Export population data to Excel with formatting."""
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Population", index=False)

            # Get worksheet
            worksheet = writer.sheets["Population"]

            # Apply header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (AttributeError, TypeError):
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Add summary sheet
            summary_data = self._generate_population_summary(df)
            summary_df = pd.DataFrame(list(summary_data.items()), columns=["Metric", "Value"])
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

    def _export_excel_simulation(self, simulation_data, filepath):
        """Export simulation results to Excel with multiple sheets."""
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # Main inequality metrics sheet
            if isinstance(simulation_data, dict):
                if "inequality_metrics" in simulation_data:
                    simulation_data["inequality_metrics"].to_excel(writer, sheet_name="Inequality_Metrics", index=False)

                # Additional sheets for other data
                for key, data in simulation_data.items():
                    if key != "inequality_metrics" and isinstance(data, pd.DataFrame):
                        sheet_name = key.replace("_", " ").title()[:31]  # Excel sheet name limit
                        data.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                simulation_data.to_excel(writer, sheet_name="Results", index=False)

            # Apply formatting to all sheets
            for sheet_name in writer.sheets:
                self._format_excel_sheet(writer.sheets[sheet_name])

    def _export_excel_analysis(self, population_data, simulation_results, analysis_summary, filepath):
        """Export comprehensive analysis to Excel."""
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # Executive Summary sheet
            summary_df = pd.DataFrame(list(analysis_summary.items()), columns=["Metric", "Value"])
            summary_df.to_excel(writer, sheet_name="Executive_Summary", index=False)

            # Population data
            if isinstance(population_data, list):
                pop_df = pd.DataFrame(population_data)
            else:
                pop_df = population_data.copy()
            pop_df.to_excel(writer, sheet_name="Population_Data", index=False)

            # Simulation results
            if isinstance(simulation_results, dict):
                for key, data in simulation_results.items():
                    if isinstance(data, pd.DataFrame):
                        sheet_name = key.replace("_", " ").title()[:31]
                        data.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                simulation_results.to_excel(writer, sheet_name="Simulation_Results", index=False)

            # Apply formatting
            for sheet_name in writer.sheets:
                self._format_excel_sheet(writer.sheets[sheet_name])

    def _export_json_population(self, df, filepath):
        """Export population data to JSON with metadata."""
        export_data = {
            "metadata": {
                "export_timestamp": datetime.now().isoformat(),
                "total_employees": len(df),
                "data_columns": list(df.columns),
                "export_version": "1.0",
            },
            "summary_statistics": self._generate_population_summary(df),
            "employee_data": df.to_dict("records"),
        }

        with open(filepath, "w") as f:
            json.dump(export_data, f, indent=2, default=str)

    def _export_json_simulation(self, simulation_data, filepath):
        """Export simulation results to JSON."""
        if isinstance(simulation_data, dict):
            export_data = simulation_data.copy()
            # Convert DataFrames to records
            for key, data in export_data.items():
                if isinstance(data, pd.DataFrame):
                    export_data[key] = data.to_dict("records")
        else:
            export_data = {"simulation_results": simulation_data.to_dict("records")}

        export_data["metadata"] = {"export_timestamp": datetime.now().isoformat(), "export_version": "1.0"}

        with open(filepath, "w") as f:
            json.dump(export_data, f, indent=2, default=str)

    def _export_json_analysis(self, population_data, simulation_results, analysis_summary, filepath):
        """Export comprehensive analysis to JSON."""
        export_data = {
            "metadata": {
                "export_timestamp": datetime.now().isoformat(),
                "export_version": "1.0",
                "analysis_type": "comprehensive",
            },
            "executive_summary": analysis_summary,
            "population_data": (
                population_data if isinstance(population_data, list) else population_data.to_dict("records")
            ),
            "simulation_results": simulation_results,
        }

        # Convert DataFrames in simulation_results if needed
        if isinstance(simulation_results, dict):
            for key, data in export_data["simulation_results"].items():
                if isinstance(data, pd.DataFrame):
                    export_data["simulation_results"][key] = data.to_dict("records")

        with open(filepath, "w") as f:
            json.dump(export_data, f, indent=2, default=str)

    def _format_excel_sheet(self, worksheet):
        """Apply consistent formatting to Excel worksheet."""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Format header row
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, TypeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _generate_population_summary(self, df):
        """Generate summary statistics for population data."""
        summary = {
            "Total Employees": len(df),
            "Average Salary": f"£{df['salary'].mean():.2f}" if "salary" in df.columns else "N/A",
            "Median Salary": f"£{df['salary'].median():.2f}" if "salary" in df.columns else "N/A",
            "Salary Std Dev": f"£{df['salary'].std():.2f}" if "salary" in df.columns else "N/A",
        }

        if "gender" in df.columns:
            gender_counts = df["gender"].value_counts()
            for gender, count in gender_counts.items():
                summary[f"{gender.title()} Employees"] = f"{count} ({count/len(df)*100:.1f}%)"

        if "level" in df.columns:
            level_counts = df["level"].value_counts().sort_index()
            for level, count in level_counts.items():
                summary[f"Level {level} Employees"] = f"{count} ({count/len(df)*100:.1f}%)"

        return summary

    def _generate_analysis_summary(self, population_data, simulation_results):
        """Generate executive summary for comprehensive analysis."""
        summary = {
            "Analysis Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Population Size": len(population_data) if isinstance(population_data, list) else len(population_data),
        }

        # Add simulation-specific metrics if available
        if isinstance(simulation_results, dict) and "inequality_metrics" in simulation_results:
            metrics_df = simulation_results["inequality_metrics"]
            if not metrics_df.empty:
                initial_gini = metrics_df.iloc[0]["gini_coefficient"]
                final_gini = metrics_df.iloc[-1]["gini_coefficient"]
                summary["Initial Gini Coefficient"] = f"{initial_gini:.4f}"
                summary["Final Gini Coefficient"] = f"{final_gini:.4f}"
                summary["Gini Improvement"] = f"{(initial_gini - final_gini):.4f}"
                summary["Total Review Cycles"] = len(metrics_df) - 1

        return summary


def main():
    """Command-line interface for data export system."""
    parser = argparse.ArgumentParser(description="Export employee simulation data")
    parser.add_argument("--population-file", required=True, help="Path to population data file")
    parser.add_argument("--simulation-file", help="Path to simulation results file")
    parser.add_argument(
        "--formats",
        nargs="+",
        choices=["csv", "excel", "json"],
        default=["csv", "excel", "json"],
        help="Export formats",
    )
    parser.add_argument("--output-dir", default="artifacts", help="Output directory")
    parser.add_argument("--analysis-report", action="store_true", help="Generate comprehensive analysis report")

    args = parser.parse_args()

    # Initialize export system
    exporter = DataExportSystem(output_dir=args.output_dir)

    # Load population data
    try:
        if args.population_file.endswith(".json"):
            with open(args.population_file, "r") as f:
                pop_data = json.load(f)
                if isinstance(pop_data, dict) and "employee_data" in pop_data:
                    population_data = pop_data["employee_data"]
                else:
                    population_data = pop_data
        elif args.population_file.endswith(".csv"):
            population_data = pd.read_csv(args.population_file)
        else:
            LOGGER.error(f"Unsupported population file format: {args.population_file}")
            return

        LOGGER.info(f"Loaded population data from {args.population_file}")
    except Exception as e:
        LOGGER.error(f"Failed to load population data: {e}")
        return

    # Load simulation data if provided
    simulation_data = None
    if args.simulation_file:
        try:
            if args.simulation_file.endswith(".json"):
                with open(args.simulation_file, "r") as f:
                    simulation_data = json.load(f)
            elif args.simulation_file.endswith(".csv"):
                simulation_data = pd.read_csv(args.simulation_file)

            LOGGER.info(f"Loaded simulation data from {args.simulation_file}")
        except Exception as e:
            LOGGER.error(f"Failed to load simulation data: {e}")

    # Export data based on requirements
    if args.analysis_report and simulation_data is not None:
        # Generate comprehensive analysis report
        exported_files = exporter.export_analysis_report(population_data, simulation_data, format_types=args.formats)
    else:
        # Export individual datasets
        exported_files = exporter.export_employee_population(population_data, format_types=args.formats)

        if simulation_data is not None:
            sim_files = exporter.export_simulation_results(simulation_data, format_types=args.formats)
            exported_files.update(sim_files)

    # Report exported files
    print("\nExported files:")
    for format_type, filepath in exported_files.items():
        print(f"  {format_type.upper()}: {filepath}")


if __name__ == "__main__":
    main()
